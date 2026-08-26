#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
内容策划生成器（社媒运营）—— 现已升级为「任意 OpenAI 兼容大模型」通用客户端
- 默认智谱 GLM（GLM-4.7-Flash），也可通过 .env 换成任意 OpenAI 兼容模型：
  LLM_API_KEY=sk-xxx
  LLM_BASE_URL=https://api.deepseek.com/v1        # 或 Kimi/通义/豆包/GLM/本地 vLLM 等兼容地址
  LLM_MODEL=deepseek-chat                          # 对应厂商的模型名
- 只填 ZHIPU_API_KEY 时行为不变（默认走智谱）
- 可选：读取「社媒运营中台_飞书模板.xlsx」的 ⑥本人 / ⑦竞品 / ⑧看板 做差距分析
依赖：标准库 + openpyxl（中控分析用，脚本已带容错）
"""
import urllib.request, json, os, sys, time

# ---------- 加载 .env（不依赖第三方库） ----------
def load_env(path=".env"):
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

load_env()

# ---------- LLM 配置（兼容任意 OpenAI 格式厂商） ----------
# 优先用 LLM_* 通用变量；没填则回退智谱（ZHIPU_API_KEY）
LLM_API_KEY = os.environ.get("LLM_API_KEY") or os.environ.get("ZHIPU_API_KEY")
LLM_BASE_URL = os.environ.get("LLM_BASE_URL", "https://open.bigmodel.cn/api/paas/v4")
LLM_MODEL = os.environ.get("LLM_MODEL", "GLM-4.7-Flash")
# base_url 归一化：兼容「带 /chat/completions」与「只到 /v1」两种填法
BASE = LLM_BASE_URL.rstrip("/")
CHAT_URL = BASE if BASE.endswith("/chat/completions") else BASE + "/chat/completions"

KEY = LLM_API_KEY            # 兼容旧引用（digest_competitor.py 等仍用 glm.KEY）
URL = CHAT_URL               # 兼容旧引用（digest_competitor.py 等仍用 glm.URL）
USER_MODEL = LLM_MODEL       # 兼容旧引用
LAST_MODEL = USER_MODEL      # 实际成功调用的模型名（被限流时会回退）

PLACEHOLDER_HITS = ["把你的key", "sk-把你的", "your-key", "sk-your", "your_api", "sk-xxxx"]


def _key_is_placeholder(key):
    """检测是否仍是占位符（用户忘了填真实 key）。"""
    kl = (key or "").strip().lower()
    return not kl or any(h in kl for h in PLACEHOLDER_HITS)


def chat(messages, model=None, temperature=0.85, timeout=90):
    """调用任意 OpenAI 兼容大模型。默认用 .env 配置；显式传 model 时优先用它。
    模型名不存在/被限流(400/404/429)时自动回退到候选列表；弱网自动重试。"""
    if _key_is_placeholder(KEY):
        print("-" * 60, flush=True)
        print("[ERROR] 未配置有效的大模型 API Key（检测到占位符或为空）。", flush=True)
        print("        请编辑 .env，填入真实的三件套：", flush=True)
        print("        LLM_API_KEY=你的真实key", flush=True)
        print("        LLM_BASE_URL=https://api.deepseek.com/v1   （或其它 OpenAI 兼容地址）", flush=True)
        print("        LLM_MODEL=deepseek-chat", flush=True)
        print("        复制模板：copy .env.example .env  然后编辑。", flush=True)
        print("-" * 60, flush=True)
        raise RuntimeError("未配置有效的 LLM API Key（占位符/为空）。请先填入真实 key。")
    model = model or LLM_MODEL
    # 候选回退列表：默认 GLM 系列兜底；用户显式自定义模型时只用它自己的
    if os.environ.get("LLM_MODEL"):
        candidates = [model, "deepseek-chat", "moonshot-v1-8k", "qwen-plus"]  # 常见厂商候选
    else:
        candidates = [model, "glm-4-flash", "glm-4-plus", "glm-4"]
    tried = []
    for m in candidates:
        if m in tried:
            continue
        tried.append(m)
        payload = json.dumps({
            "model": m, "messages": messages,
            "temperature": temperature, "stream": False
        }).encode("utf-8")
        req = urllib.request.Request(CHAT_URL, data=payload, headers={
            "Authorization": f"Bearer {KEY}",
            "Content-Type": "application/json"
        })
        for attempt in range(3):
            try:
                with urllib.request.urlopen(req, timeout=timeout) as r:
                    global LAST_MODEL
                    LAST_MODEL = m
                    return json.loads(r.read().decode("utf-8"))
            except urllib.error.HTTPError as e:
                body = e.read().decode("utf-8", "ignore")
                if e.code in (400, 404, 429):
                    break  # 模型名不存在或被限流，试下一个候选模型
                raise
            except Exception:
                if attempt < 2:
                    time.sleep(2 * (attempt + 1))  # 弱网抖动重试
                    continue
                raise
    raise RuntimeError(f"所有模型均失败，已尝试: {tried}")

def reply_text(resp):
    return resp["choices"][0]["message"]["content"]

# ---------- 1. 连通测试 ----------
def test_key():
    print(">> [1/3] 测试 GLM API 连通性 ...")
    r = chat([{"role": "user", "content": "只回复四个字：API连通成功"}], temperature=0.1)
    print("   模型返回:", reply_text(r).strip())
    print("   Key 有效，可正常使用。\n")

# ---------- 2. 内容策划生成 ----------
# 账号定位从 .env 读取（ACCOUNT_* 环境变量），兜底为完全中性的通用描述
ACCOUNT_CTX = f"""你是资深的小红书/抖音内容策划。
我们有两个账号需要策划：
【账号A：个人IP（小红书）】
- 人设：{os.environ.get('ACCOUNT_PERSONA', '垂直领域内容创作者，有真实经历')}
- 受众：{os.environ.get('ACCOUNT_AUDIENCE', '关注该领域的普通用户 / 学习者')}
- 内容调性：真实经历+干货、有温度、建立信任（种草个人）
- 形式：初期以「配图+文案」图文为主

【账号B：机构号（抖音）】
- 人设：由专业人士创立的垂直领域机构
- 受众：同上，但更偏「机构权威/结果导向」
- 内容调性：专业口播、学员案例、直播切片，给个人号导流
- 形式：短视频口播/混剪，后期可配音"""

def generate_plan():
    print(">> [2/3] 生成「个人IP / 机构号」内容策划 ...")
    user_prompt = """请基于上面两个账号的定位，各产出 5 条可立即执行的爆款选题（共10条）。
每条按如下结构输出（用 markdown）：
## 账号A：个人IP（小红书）
### 选题1：<标题>
- 类型：图文/短视频
- 钩子（前3秒/首句）：
- 正文要点（3-5条）：
- 话题标签：#xxx #xxx
- 建议发布时间：

（账号B 同理，标题写「账号B：机构号（抖音）」）

要求：选题贴合目标领域真实痛点，有数据感或反差感，避免空泛。"""
    r = chat([
        {"role": "system", "content": ACCOUNT_CTX},
        {"role": "user", "content": user_prompt}
    ], temperature=0.9)
    text = reply_text(r)
    with open("GLM内容策划_个人IP_机构号.md", "w", encoding="utf-8") as f:
        f.write("# GLM 生成的内容策划（个人IP / 机构号）\n\n")
        f.write(f"> 实际调用模型：{LAST_MODEL}（主模型 {LLM_MODEL} 被限流时自动回退）  |  生成方式：API 自动调用\n\n")
        f.write(text)
    print("   已保存：GLM内容策划_个人IP_机构号.md\n")
    return text

# ---------- 3. 中控差距分析（可选） ----------
def analyze_zhongkong():
    xlsx = os.environ.get("ZHONGKONG_XLSX", "社媒运营中台_飞书模板.xlsx")
    out_md = os.environ.get("ZHONGKONG_OUT", "GLM中控差距分析.md")
    if not os.path.exists(xlsx):
        print(">> [3/3] 未找到中控 xlsx，跳过差距分析。\n")
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx, data_only=True)
    except Exception as e:
        print(f">> [3/3] 读取中控失败：{e}\n")
        return
    def sheet_rows(sheet, names):
        if sheet not in wb.sheetnames:
            return None
        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        try:
            idx = {name: (headers.index(name) + 1) for name in names}
        except ValueError:
            return None
        rows = []
        for row in ws.iter_rows(min_row=2):
            vals = {name: row[idx[name] - 1].value for name in names}
            if all(v is None for v in vals.values()):
                continue
            rows.append(vals)
        return rows
    self_rows = sheet_rows("⑥本人账号监控", ["播放量", "点赞数", "收藏数", "评论数"])
    comp_rows = sheet_rows("⑦竞品监控", ["播放量", "点赞数", "收藏数", "评论数"])
    if not self_rows or not comp_rows:
        print(">> [3/3] 中控表结构不匹配或为空，跳过。\n")
        return
    def agg(rows):
        n = len(rows)
        def m(k):
            return sum(r[k] or 0 for r in rows) / n
        play, digg, col, com = m("播放量"), m("点赞数"), m("收藏数"), m("评论数")
        denom = max(1, sum(1 for r in rows if (r["播放量"] or 0) > 0))
        def rate(num_key):
            return sum((r[num_key] or 0) / (r["播放量"] or 1) for r in rows if (r["播放量"] or 0) > 0) / denom
        return {"播放量": play, "点赞数": digg, "收藏数": col, "评论数": com,
                "点赞率": rate("点赞数"), "收藏率": rate("收藏数"), "评论率": rate("评论数")}
    s = agg(self_rows)
    c = agg(comp_rows)
    def pct(x):
        return f"{x * 100:.1f}%"
    prompt = f"""这是我们两个账号的中控数据（本人=个人IP，竞品=对标账号均值）：
【绝对量均值】本人 播放{s['播放量']:.0f}/点赞{s['点赞数']:.0f}/收藏{s['收藏数']:.0f}/评论{s['评论数']:.0f}
           竞品 播放{c['播放量']:.0f}/点赞{c['点赞数']:.0f}/收藏{c['收藏数']:.0f}/评论{c['评论数']:.0f}
【互动率均值】本人 点赞率{pct(s['点赞率'])}/收藏率{pct(s['收藏率'])}/评论率{pct(s['评论率'])}
           竞品 点赞率{pct(c['点赞率'])}/收藏率{pct(c['收藏率'])}/评论率{pct(c['评论率'])}
请做一份差距分析，指出我们在哪些指标落后/领先，并给出 3-5 条可执行的「下一步内容方向」建议（结合目标领域赛道）。用 markdown 小标题分点。"""
    print(">> [3/3] 基于中控做差距分析 ...")
    r = chat([{"role": "user", "content": prompt}], temperature=0.5)
    text = reply_text(r)
    with open(out_md, "w", encoding="utf-8") as f:
        f.write(f"# GLM 中控差距分析\n\n> 实际调用模型：{LAST_MODEL}  |  数据来源：{xlsx}\n\n")
        f.write(text)
    print(f"   已保存：{out_md}\n")

if __name__ == "__main__":
    if _key_is_placeholder(KEY):
        print("ERROR: 未配置有效的大模型 API Key（占位符/为空）。请在 .env 中写入 LLM_API_KEY=你的真实key（或 ZHIPU_API_KEY）。")
        sys.exit(1)
    test_key()
    generate_plan()
    analyze_zhongkong()
    print("全部完成")
