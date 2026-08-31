# -*- coding: utf-8 -*-
"""
analytics.py —— 企业级热度引擎 + 验证管线（极光AIGC · 免费本地版）

设计依据：research/企业级竞品爬取与选题方案_2026-08-31.md
- 不依赖任何付费工具（千瓜/新红/灰豚），全部本地可复现。
- 不调用 GLM/智谱，纯数值计算 + 规则，可离线跑、可测试。

对外能力：
  1) 归一化：build_note() 把 MediaCrawler CSV 行 / 竞品 xlsx 行统一成 canonical 字段。
  2) 验证管线：
     - validate_completeness() 字段完整性校验（缺必填即标异常）
     - detect_brush()          反刷量特征识别（假数据风险分）
     - relative_perf()         相对表现 R + 三角校验（极端值标人工复核）
  3) 热度引擎：
     - compute_baseline()      同赛道基线（中位数/均值，对数友好）
     - score_note()            单篇：CES_local / 互动率 / 收藏率 / R / 增速G / 热度分0-100
     - score_notes()           批量打分，输出可直接进飞书「热度看板」
  4) 真实 vs 虚荣指标说明（real_metrics_summary）。

CLI（便于本地验证）：
  python analytics.py --csv tools/MediaCrawler/data/xhs/csv/search_contents_2026-08-25.csv --top 20
  python analytics.py --xlsx pipeline/竞品选题库_合并.xlsx --top 10
"""
import os, sys, json, math, csv, argparse
from datetime import datetime, timezone
from statistics import median
from openpyxl import Workbook

# ---------- canonical 字段 ----------
# note_id, title, account, url, publish_time(str), liked, collected, comment, share,
# desc, tags(list), source, last_modify_ts
CES_WEIGHTS = {  # 对标 2026 小红书 CES 官方权重（星云5.0）
    "liked": 1, "collected": 1, "comment": 4, "share": 4,
    # 以下字段本地无，需千瓜/蒲公英校准（方案 2.3 第6条局限标注）：
    # "follow": 8, "screenshot_save": 6, "deep_read": 4
}
AVAILABLE_CES = list(CES_WEIGHTS.keys())

# 反刷量阈值（方案 1.3.2）
COMMENT_RATE_SUSPICIOUS = 0.005   # 评论率 < 0.5% 且高赞 -> 无真实讨论
COLLECT_LIKE_RATIO_SUSPICIOUS = 3.0  # 收藏/赞 > 3 -> 疑似刷藏
R_EXTREME = 50.0  # 相对表现 R > 50x -> 极端爆款/疑似刷量，需人工复核

# 热度分权重（方案 2.3.5）：CES 0.7 + 增速 0.3
W_CES = 0.7
W_GROWTH = 0.3

# 字段完整性必填
REQUIRED_FIELDS = ["note_id", "title", "liked", "collected", "comment", "publish_time", "account"]


def _coerce_int(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip().replace(",", "").replace("+", "")
    if s.endswith("万"):
        try:
            return int(float(s[:-1]) * 10000)
        except ValueError:
            return 0
    if s.endswith("w") or s.endswith("W"):
        try:
            return int(float(s[:-1]) * 10000)
        except ValueError:
            return 0
    s = "".join(ch for ch in s if ch.isdigit() or ch == ".")
    try:
        return int(float(s)) if s else 0
    except ValueError:
        return 0


def _parse_time(s):
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    # 可能是毫秒/秒时间戳
    try:
        v = float(s)
        if v > 1e12:
            v = v / 1000
        return datetime.fromtimestamp(v)
    except ValueError:
        return None


def _clean_url(u):
    if not u:
        return ""
    return u.split("?")[0]  # 截断 ?xsec_token=... 飞书拒收


def _parse_tags(t):
    if isinstance(t, list):
        return [str(x).strip() for x in t if str(x).strip()]
    if not t:
        return []
    return [x.strip() for x in str(t).replace("，", ",").split(",") if x.strip()]


def build_note(raw):
    """把任意来源行归一化为 canonical note dict。raw 支持两种命名：
    MediaCrawler CSV：note_id/title/nickname/liked_count/.../tag_list/note_url/time/source_keyword
    竞品 xlsx：笔记ID/标题/账号/点赞数/.../话题标签/笔记链接/发布日期/数据来源
    """
    g = lambda *keys: next((raw.get(k) for k in keys if raw.get(k) not in (None, "")), None)
    note_id = g("note_id", "笔记ID")
    title = g("title", "标题")
    account = g("nickname", "账号")
    liked = _coerce_int(g("liked_count", "点赞数"))
    collected = _coerce_int(g("collected_count", "收藏数"))
    comment = _coerce_int(g("comment_count", "评论数"))
    share = _coerce_int(g("share_count", "分享数"))
    desc = g("desc", "正文", "note_desc")
    tags = _parse_tags(g("tag_list", "话题标签"))
    url = _clean_url(g("note_url", "笔记链接"))
    publish_time = g("time", "publish_time", "发布日期")
    source = g("source_keyword", "数据来源", "source")
    last_modify_ts = g("last_modify_ts", "last_modify_ts")
    return {
        "note_id": note_id,
        "title": title or "",
        "account": account or "",
        "url": url or "",
        "publish_time": publish_time or "",
        "liked": liked,
        "collected": collected,
        "comment": comment,
        "share": share,
        "desc": desc or "",
        "tags": tags,
        "source": source or "",
        "last_modify_ts": last_modify_ts,
    }


# ======================= 验证管线 =======================

def validate_completeness(note):
    """字段完整性校验（方案 1.3.1）。返回缺失字段列表；空=完整。"""
    missing = [f for f in REQUIRED_FIELDS if not note.get(f)]
    # 互动量全为 0 也视为不完整（没采到数）
    if note.get("liked", 0) == 0 and note.get("collected", 0) == 0 and note.get("comment", 0) == 0:
        if "interactions" not in missing:
            missing.append("interactions(全0)")
    return missing


def detect_brush(note, baseline=None):
    """反刷量特征识别（方案 1.3.2）。返回 (risk_score 0-100, reasons[])。
    本地仅能检查「无官方主页访问率/模板评论」的可用子集，已在 reasons 标注局限。
    """
    reasons = []
    risk = 0
    liked = note.get("liked", 0)
    collected = note.get("collected", 0)
    comment = note.get("comment", 0)
    share = note.get("share", 0)
    total = liked + collected + comment + share

    if liked > 0:
        comment_rate = comment / liked
        if comment_rate < COMMENT_RATE_SUSPICIOUS and liked >= 1000:
            risk += 35
            reasons.append(f"评论率{comment_rate:.2%}过低(<0.5%)且高赞，疑似无真实讨论")
        if collected / liked > COLLECT_LIKE_RATIO_SUSPICIOUS and liked >= 1000:
            risk += 25
            reasons.append(f"收藏/赞比{collected/liked:.1f}异常高(>3)，疑似刷藏")
    # 赞藏几乎同时极端（本地无时间序列，用赞≈藏且都巨大近似）
    if liked >= 5000 and abs(liked - collected) / max(liked, 1) < 0.02:
        risk += 15
        reasons.append("赞≈藏且量级极大，需结合时间序列复核是否为同步刷量")
    if total == 0:
        risk += 10
        reasons.append("互动全为0，可能是抓取失败/死链")
    if not reasons:
        reasons.append("未命中已知刷量特征（注：无主页访问率/模板评论字段，需官方口径二次校验）")
    return min(100, risk), reasons


def compute_baseline(notes):
    """同赛道基线：各互动量中位数 + 均值（对数归一化友好）。"""
    def col(key):
        return [n.get(key, 0) for n in notes if n.get(key, 0) > 0]
    base = {}
    for k in AVAILABLE_CES:
        vals = col(k)
        base[k] = {
            "median": median(vals) if vals else 0,
            "mean": sum(vals) / len(vals) if vals else 0,
        }
    # 总互动基线（用于相对表现 R）
    totals = [sum(n.get(k, 0) for k in AVAILABLE_CES) for n in notes if sum(n.get(k, 0) for k in AVAILABLE_CES) > 0]
    base["_total"] = {"median": median(totals) if totals else 0, "mean": sum(totals) / len(totals) if totals else 0}
    return base


def relative_perf(note, baseline):
    """相对表现 R（方案 1.3.4 / 2.3）：单篇 vs 同赛道基线。返回 dict{R_*, R_total, flag}。"""
    R = {}
    for k in AVAILABLE_CES:
        b = baseline.get(k, {}).get("median", 0) or 1
        R[f"R_{k}"] = round(note.get(k, 0) / b, 2)
    total = sum(note.get(k, 0) for k in AVAILABLE_CES)
    b_total = baseline.get("_total", {}).get("median", 0) or 1
    R_total = round(total / b_total, 2)
    flag = ""
    if R_total > R_EXTREME:
        flag = "极端爆款/疑似刷量，需人工复核（三角校验：与千瓜/蒲公英口径对不上时以官方为准）"
    return {"R": R, "R_total": R_total, "flag": flag}


# ======================= 热度引擎 =======================

def _percentile_rank(values):
    """返回每个值在排序中的百分位 (0-100)。"""
    s = sorted(values)
    n = len(s)
    out = []
    for v in values:
        # 小于等于 v 的个数 / n
        le = sum(1 for x in s if x <= v)
        out.append(round(100 * le / n, 1))
    return out


def score_note(note, baseline):
    """单篇热度打分（方案 2.3）。返回扩展后的 note dict。"""
    # 1) 相对 CES（只算可得字段，权重对标官方）
    ces_local = 0.0
    for k in AVAILABLE_CES:
        b = baseline.get(k, {}).get("median", 0) or 1
        Rk = note.get(k, 0) / b
        ces_local += Rk * CES_WEIGHTS[k]
    note["ces_local"] = round(ces_local, 3)

    # 2) 真实指标子集（曝光/主页访客本地无，按方案标注局限）
    total = sum(note.get(k, 0) for k in AVAILABLE_CES)
    note["interactions"] = total
    note["collect_rate"] = round(note["collected"] / note["liked"], 4) if note["liked"] else 0
    note["comment_rate"] = round(note["comment"] / max(note["liked"], 1), 4)
    note["share_rate"] = round(note["share"] / max(note["liked"], 1), 4)
    note["save_rate"] = round(note["collected"] / max(total, 1), 4)  # 收藏/总互动，近似"有用"信号

    # 3) 增速 G —— 本地无时间序列，用「互动速度 = 总互动/发布天数」作代理（方案标注局限）
    pt = _parse_time(note.get("publish_time"))
    now = datetime.now()
    days = max(1, (now - pt).days) if pt else 30
    note["age_days"] = days
    velocity = total / days
    note["velocity_per_day"] = round(velocity, 1)

    # 4) 相对表现 R
    rp = relative_perf(note, baseline)
    note["R_total"] = rp["R_total"]
    note["R_detail"] = rp["R"]
    note["r_flag"] = rp["flag"]

    # 5) 热度分 = 百分位排名融合（CES 0.7 + 增速 0.3），0-100
    note["_ces_raw"] = ces_local
    note["_vel_raw"] = velocity
    return note


def score_notes(notes, baseline=None):
    """批量打分。返回带 heat_score 的 notes（按热度降序）。"""
    # 去重：同 note_id 保留互动总量最大的一条（同一篇可能被多个搜索词命中）
    _best = {}
    for n in notes:
        nid = n.get("note_id")
        if not nid:
            continue
        tot = sum(n.get(k, 0) for k in AVAILABLE_CES)
        if nid not in _best or tot > _best[nid]["_tot"]:
            _best[nid] = {"note": n, "_tot": tot}
    notes = [v["note"] for v in _best.values()]
    notes = [n for n in notes if validate_completeness(n) == []]  # 先过滤不完整
    if baseline is None:
        baseline = compute_baseline(notes)
    scored = [score_note(n, baseline) for n in notes]
    # 百分位排名
    ces_pct = _percentile_rank([n["_ces_raw"] for n in scored])
    vel_pct = _percentile_rank([n["_vel_raw"] for n in scored])
    for i, n in enumerate(scored):
        heat = W_CES * (ces_pct[i] / 100) + W_GROWTH * (vel_pct[i] / 100)
        n["heat_score"] = round(heat * 100, 1)
        n["ces_pct"] = ces_pct[i]
        n["vel_pct"] = vel_pct[i]
        # 验证挂到每篇
        miss = validate_completeness(n)
        n["completeness_missing"] = miss
        risk, reasons = detect_brush(n, baseline)
        n["brush_risk"] = risk
        n["brush_reasons"] = reasons
    scored.sort(key=lambda x: x["heat_score"], reverse=True)
    for i, n in enumerate(scored, 1):
        n["heat_rank"] = i
    return scored, baseline


def real_metrics_summary(notes):
    """真实 vs 虚荣指标（方案 2.2）。返回人话总结，标注本地不可得项。"""
    if not notes:
        return "无数据"
    heats = [n["heat_score"] for n in notes]
    brush_high = [n for n in notes if n.get("brush_risk", 0) >= 50]
    collect_rates = [n["collect_rate"] for n in notes if n["liked"]]
    avg_cr = sum(collect_rates) / len(collect_rates) if collect_rates else 0
    return {
        "笔记数": len(notes),
        "热度分区间": f"{min(heats):.0f}–{max(heats):.0f}",
        "平均收藏率(藏/赞)": f"{avg_cr:.1%}",
        "高刷量风险(≥50分)": f"{len(brush_high)} 篇（需人工复核）",
        "本地不可得(需官方/付费校准)": "曝光量、主页访客、自然流量占比、截图保存、达人影响力——方案2.3第6条已标注低估局限",
        "判断口径": "热度分=CES_local(0.7)+互动速度(0.3) 的百分位融合；看相对R不看绝对值；增速用互动速度代理",
    }


# ======================= IO =======================

def load_notes_from_csv(path):
    out = []
    with open(path, encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            out.append(build_note(row))
    return out


def load_notes_from_xlsx(path):
    from openpyxl import load_workbook
    out = []
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r):
            continue
        out.append(build_note({headers[i]: r[i] for i in range(len(headers))}))
    return out


HEAT_BOARD_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "热度看板.xlsx")
HEAT_HEADERS = ["排名", "标题", "笔记ID", "账号", "点赞数", "收藏数", "评论数", "分享数",
                "热度分", "相对表现R", "互动速度", "收藏率", "评论率", "状态"]


def write_heat_board_xlsx(scored, path=HEAT_BOARD_XLSX):
    """把打分结果写成飞书「热度看板」数据源 xlsx（运营小白打开即知什么火）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "热度看板"
    ws.append(HEAT_HEADERS)
    for n in scored:
        ws.append([
            n.get("heat_rank", 0), n.get("title", "")[:60], n.get("note_id", ""), n.get("account", ""),
            n.get("liked", 0), n.get("collected", 0), n.get("comment", 0), n.get("share", 0),
            n.get("heat_score", 0), n.get("R_total", 0), n.get("velocity_per_day", 0),
            n.get("collect_rate", 0), n.get("comment_rate", 0),
            "估算(无官方曝光口径)" if n.get("r_flag") else "估算",
        ])
    wb.save(path)
    return path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv")
    ap.add_argument("--xlsx")
    ap.add_argument("--top", type=int, default=15)
    a = ap.parse_args()
    if a.csv:
        notes = load_notes_from_csv(a.csv)
    elif a.xlsx:
        notes = load_notes_from_xlsx(a.xlsx)
    else:
        print("ERROR: 需 --csv 或 --xlsx"); sys.exit(1)
    print(f"读入 {len(notes)} 条，过滤不完整后打分...")
    scored, base = score_notes(notes)
    print("基线(median)：", {k: base[k]["median"] for k in AVAILABLE_CES}, "总互动中位:", base["_total"]["median"])
    print(json.dumps(real_metrics_summary(scored), ensure_ascii=False, indent=2))
    print(f"\nTop {a.top}（热度分降序）：")
    for n in scored[:a.top]:
        flag = " ⚠️" + n["r_flag"] if n["r_flag"] else ""
        brush = f" 刷量风险{n['brush_risk']}" if n["brush_risk"] >= 35 else ""
        print(f"  {n['heat_rank']:>2}. [{n['heat_score']:>5}] {n['title'][:26]:<26} "
              f"赞{n['liked']}藏{n['collected']}评{n['comment']} R={n['R_total']} 速{n['velocity_per_day']}/d{brush}{flag}")


if __name__ == "__main__":
    main()
