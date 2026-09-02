#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
小红书 MVP 一键成稿 + 生图（极光AIGC · 去 GLM 版）

设计（方案 §3 + 用户硬约束「内容生成不用 GLM/智谱」）：
- 文案/配图方案【由 AI（WorkBuddy 自带模型）注入】，本脚本不调用任何外部 LLM。
- 本脚本职责：① 读入 agent 注入的内容 JSON（--inject）；② 按 provider 出图或只出提示词；
  ③ 落盘 md + json + 图片（或提示词卡片）；④ 供 push_to_feishu_content.py 推飞书。

生图（可插拔 + 自动降级，见 pipeline/image_provider.py）
- auto（默认）：探测到生图 API 就调 API；没配 key 则降级 prompt_only
- prompt_only：不出图，产出「运营提示词卡片」，由运营去豆包/即梦等免费网站生成
- openai     ：OpenAI 兼容 /images/generations（OpenAI、硅基流动、火山方舟等）
- pollinations / cogview：保留为显式选项，不再是默认（效果一般 / 无权限）

两条硬规则（企业级研究结论，详见 docs/生图配置指南.md）：
1. 配图提示词必须【锚定正文】：cover/inner_images 的 prompt 要对应正文真实提到的
   截图/体验/清单/对比，禁止「学姐侧身指着图表」式脱节场景。
2. 图片里的【文字不靠模型渲染】——中文必乱码。提示词强制含「无文字/无水印」，
   标题由运营后期用稿定/Canva 压上（免费商用字体：思源黑体 / 阿里巴巴普惠体）。

⚠️ 配图提示词必须【锚定正文】（用户反馈 Point 4）：agent 注入时，cover/inner_images 的
   prompt 必须对应正文里真实提到的截图/体验/清单/对比，禁止「学姐侧身指着图表」式脱节场景。

内容 JSON 约定（agent 注入）：
{
  "topic": "选题", "title": "标题≤20字", "hook": "钩子",
  "body": "正文（\\n分段）", "tags": ["#a","#b"], "publish_tip": "发布建议",
  "cover": {"prompt": "画面主体描述(中文,锚定正文)", "caption":"封面上大字", "layout":"版式"},
  "inner_images": [ {"prompt":"...","caption":"...","layout":"..."} x3 ]
}

注：prompt 可只写「画面里有什么」，品牌风格锁与固定字眼
    （风格统一 / 无文字 / 无水印）由 image_prompt.ensure_prompt() 自动补齐。

用法：
  （每个号先分析一次品牌锁）python pipeline/brand_analyzer.py analyze \
      --account <id> --cover-dir <该号历史封面目录> [--corpus <文案语料>]
  然后成稿：
  python xhs_mvp.py --account <id> --inject xhs_posts/xhs_<slug>.json                 # 注入内容 + 生图(auto)
  python xhs_mvp.py --account <id> --inject xhs_posts/xhs_<slug>.json --provider prompt_only  # 强制只出提示词
  python xhs_mvp.py --account <id> --inject xhs_posts/xhs_<slug>.json --no_image       # 只出文案
  python xhs_mvp.py --account <id> --from-recommend                                    # 对已注入内容的条目生图
"""
import os, sys, json, re, argparse, time, shutil, base64, urllib.request, urllib.parse, urllib.error

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, HERE)
sys.path.insert(0, ROOT)

from image_provider import (generate as provider_generate, resolve_provider,
                            describe as describe_provider)
from image_prompt import ensure_prompt, build_operator_card, validate_prompt
from brand_analyzer import brand_copy_block

# 生图开关（main 中按 --no_image / --size / --provider 覆盖）
NO_IMAGE = False
IMG_SIZE = "1080x1440"  # 小红书 3:4 竖版（信息流展示面积最大，原 1024x1024 方图不符合规范）
# auto：探测到生图 API 就用 API，否则降级 prompt_only（只出提示词给运营去免费网站）
PROVIDER = os.environ.get("IMAGE_PROVIDER", "auto")


def gen_image(prompt, size=IMG_SIZE, retries=3, provider=None):
    """统一走 image_provider（可插拔 + 自动降级）。prompt_only 时返回 (None, None)。"""
    return provider_generate(prompt, size=size, provider=provider or PROVIDER, retries=retries)


# 注：gen_image_cogview / gen_image_pollinations 已迁移到 image_provider.py，
#     统一由 gen_image() -> provider_generate() 分发（可插拔 + auto 降级）。


def download(url, path, retries=3):
    if not url.startswith(("http://", "https://")):
        shutil.copy(url, path)
        return True
    last = None
    for i in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=60) as r, open(path, "wb") as f:
                f.write(r.read())
            return True
        except Exception as e:
            last = e
            time.sleep(2)
    raise RuntimeError(f"下载图片失败：{last}")


def gen_images_for_post(imgs, img_dir, size=IMG_SIZE, provider=None, account=None):
    """
    对 封面+内页 生图并落盘。
    - 每张 prompt 先过 ensure_prompt()：补品牌风格锁 + 强制固定字眼（风格统一/无文字/无水印）
    - prompt_only 模式直接返回 (None, None)：不出图，改由调用方生成运营提示词卡片
    - account：账号 ID，用于加载该号专属品牌锁（图提示词风格隔离）
    """
    p = resolve_provider(provider or PROVIDER)
    if p == "prompt_only":
        return None, None

    os.makedirs(img_dir, exist_ok=True)
    saved, urls = {}, {}
    jobs = [("cover", imgs.get("cover") or {}, "封面")]
    jobs += [(f"inner{i}", im, f"内页{i}")
             for i, im in enumerate(imgs.get("inner_images") or [], 1)]

    for key, im, role in jobs:
        # 兼容旧 schema：V6 用 subject 存图描述，V7 改 prompt。两者皆无则告警，绝不静默跳过
        raw = im.get("prompt") or im.get("subject") or ""
        if not raw:
            print(f"   ⚠️ 跳过「{role}」：既无 prompt 也无 subject"
                  f"（旧 JSON 请迁移为 prompt，或补 subject 字段）")
            continue
        prompt = ensure_prompt(raw, role=role, account=account)
        f, u = gen_image(prompt, size=size, provider=p)
        out = os.path.join(img_dir, f"{key}.png")
        download(f, out)
        saved[key] = out
        urls[key] = u
    return saved, urls


def slugify(topic):
    return re.sub(r"[^\w一-鿿]+", "_", topic)[:30]


def load_recommend_topics(limit=0):
    from openpyxl import load_workbook
    p = os.path.join(HERE, "今日选题推荐.xlsx")
    if not os.path.exists(p):
        return []
    wb = load_workbook(p, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    topics = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r):
            continue
        t = r[idx["选题标题"]]
        if t:
            topics.append(str(t).strip())
    return topics[:limit] if limit else topics


def gen_post_from_data(data, out_dir, provider=None, account=None):
    """agent 已注入完整内容 JSON：生图 + 落盘 md/json/images。

    account：账号 ID，用于加载该号专属品牌锁（文案风格锁注入成稿 + 图提示词隔离）。
    """
    topic = data.get("topic") or data.get("title") or "未命名选题"
    slug = slugify(topic)
    md_path = os.path.join(out_dir, f"xhs_{slug}.md")
    json_path = os.path.join(out_dir, f"xhs_{slug}.json")
    img_path = os.path.join(out_dir, f"xhs_{slug}_images.json")
    os.makedirs(out_dir, exist_ok=True)

    # P2：注入本号文案风格锁（账号专属口吻约束，WorkBuddy 模型生成正文时遵守）
    copy_lock_md = ""
    if account:
        try:
            from core import accounts as _accts
            _brand = _accts.load_brand(account)
            _copy = (_brand or {}).get("copy") or {}
            if _copy:
                copy_lock_md = "\n## 本号文案风格锁\n\n" + brand_copy_block(_copy) + "\n"
        except Exception:
            copy_lock_md = ""

    with open(md_path, "w", encoding="utf-8") as f:
        f.write(f"# 小红书图文帖\n\n> 选题：{data.get('topic','')}\n\n")
        f.write(f"## 标题\n{data.get('title','')}\n\n")
        if copy_lock_md:
            f.write(copy_lock_md)
        f.write(f"## 正文\n{data.get('hook','')}\n\n{data.get('body','')}\n\n")
        f.write("## 话题标签\n" + " ".join(data.get("tags", [])) + "\n\n")
        f.write(f"## 发布建议\n{data.get('publish_tip','')}\n\n")
        f.write("## 配图方案（封面+3内页，提示词锚定正文）\n")
        # 兼容旧 schema：caption←text、prompt←subject
        c = data.get("cover", {}) or {}
        f.write(f"- **封面**：{c.get('caption') or c.get('text','')} ｜ 版式：{c.get('layout','')}\n"
                f"  - prompt: {c.get('prompt') or c.get('subject','')}\n")
        for i, im in enumerate(data.get("inner_images") or [], 1):
            f.write(f"- **内页{i}**：{im.get('caption') or im.get('text','')} ｜ 版式：{im.get('layout','')}\n"
                    f"  - prompt: {im.get('prompt') or im.get('subject','')}\n")

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    with open(img_path, "w", encoding="utf-8") as f:
        json.dump({"cover": data.get("cover", {}), "inner_images": data.get("inner_images", [])},
                  f, ensure_ascii=False, indent=2)

    image_paths, image_urls = {}, {}
    prov = resolve_provider(provider or PROVIDER)
    if not NO_IMAGE:
        if prov == "prompt_only":
            # 降级路径 A：没配生图 API —— 交付「运营提示词卡片」，而不是报错中断
            card = build_operator_card(data, account=account,
                                       mode="提示词模式（未配置生图 API）")
            card_path = os.path.join(out_dir, f"xhs_{slug}_生图提示词.md")
            with open(card_path, "w", encoding="utf-8") as f:
                f.write(card)
            data["_image_mode"] = "prompt_only"
            data["_image_card"] = card_path
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            with open(md_path, "a", encoding="utf-8") as f:
                f.write(f"\n## 配图（提示词模式）\n> 未配置生图 API，已生成运营提示词卡片：`{card_path}`\n")
            print(f"   [提示词模式] 未配置生图 API -> 已出运营提示词卡片：{os.path.basename(card_path)}")
            print("   运营复制提示词去豆包/即梦出底图，标题后期用稿定/Canva 压上。")
        else:
            print(f"   生图中（{describe_provider(prov)}）...")
            img_dir = os.path.join(out_dir, "images", slug)
            try:
                image_paths, image_urls = gen_images_for_post(
                    {"cover": data.get("cover", {}), "inner_images": data.get("inner_images", [])},
                    img_dir, size=IMG_SIZE, provider=prov, account=account)
                data["_images_local"] = image_paths
                data["_image_urls"] = image_urls
                data["_image_mode"] = prov
                with open(json_path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                with open(md_path, "a", encoding="utf-8") as f:
                    f.write("\n## 配图成品（本地路径）\n")
                    for role, p in image_paths.items():
                        f.write(f"- **{role}**：`{p}`\n")
                print(f"   配图已保存：images/{slug}/")
            except Exception as e:
                # 降级路径 B：API 调用失败 —— 自动降级为提示词卡片，保证流水线仍可交付
                print(f"   ⚠️ 生图失败：{e}")
                print("   → 自动降级为提示词卡片（流水线不中断）")
                card = build_operator_card(data, account=account,
                                           mode="提示词模式（生图 API 调用失败，已降级）")
                card_path = os.path.join(out_dir, f"xhs_{slug}_生图提示词.md")
                with open(card_path, "w", encoding="utf-8") as f:
                    f.write(card)
                data["_image_mode"] = "prompt_only_fallback"
                data["_image_card"] = card_path
                data["_image_error"] = str(e)
                with open(json_path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                print(f"   运营提示词卡片：{os.path.basename(card_path)}")

    print(f"   文案已保存：{os.path.basename(md_path)}")
    return data, md_path, img_path


def main():
    global NO_IMAGE, IMG_SIZE, PROVIDER
    ap = argparse.ArgumentParser()
    ap.add_argument("--account", required=True, help="账号 ID（= accounts/<id>/ 目录，加载该号专属品牌锁；不指定则无法继续）")
    ap.add_argument("--inject", help="注入 agent 生成的内容 JSON（必填，内容由 WorkBuddy 模型产出）")
    ap.add_argument("--from-recommend", action="store_true", help="对今日选题推荐里已注入内容的条目生图")
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--out_dir", default=os.path.join(HERE, "xhs_posts"))
    ap.add_argument("--no_image", action="store_true", help="只出文案，不生图")
    ap.add_argument("--size", default="1080x1440")
    ap.add_argument("--provider", default=None,
                    choices=["auto", "prompt_only", "openai", "pollinations", "cogview"],
                    help="生图后端：auto=有API就调/没配就降级提示词(默认)；"
                         "prompt_only=强制只出提示词给运营去免费网站；"
                         "openai=OpenAI兼容接口（支持硅基流动/火山方舟等）")
    a = ap.parse_args()
    NO_IMAGE = a.no_image
    IMG_SIZE = a.size
    if a.provider:
        PROVIDER = a.provider
    prov = resolve_provider(PROVIDER)
    print(f">> 账号：{a.account}")
    print(f">> 生图 Provider：{prov} — {describe_provider(prov)}（{IMG_SIZE}）")
    print(">> 内容来源：agent 注入（--inject）")

    # fail-fast：必须先分析出该号品牌锁，否则拿不到本号专属口吻/色板，直接报错退出
    from core import accounts as _accts
    if not os.path.exists(_accts.brand_path(a.account)):
        print(f"ERROR: 账号「{a.account}」品牌锁缺失：{_accts.brand_path(a.account)}")
        print("→ 请先分析该账号：")
        print(f"    python pipeline/brand_analyzer.py analyze --account {a.account} "
              "--cover-dir <该号历史封面目录> [--corpus <文案语料>]")
        sys.exit(2)

    if a.inject:
        if not os.path.exists(a.inject):
            print(f"ERROR: 找不到注入文件 {a.inject}")
            sys.exit(1)
        data = json.load(open(a.inject, encoding="utf-8"))
        gen_post_from_data(data, a.out_dir, provider=PROVIDER, account=a.account)
    elif a.from_recommend:
        topics = load_recommend_topics(a.limit)
        if not topics:
            print("ERROR: 今日选题推荐为空")
            sys.exit(1)
        done = 0
        for t in topics:
            slug = slugify(t)
            jp = os.path.join(a.out_dir, f"xhs_{slug}.json")
            if not os.path.exists(jp):
                print(f"   （跳过）{t[:20]}：未注入内容（请 agent 先 --inject）")
                continue
            data = json.load(open(jp, encoding="utf-8"))
            gen_post_from_data(data, a.out_dir, provider=PROVIDER, account=a.account)
            done += 1
        print(f"\n已为 {done} 条注入内容生成图文 ✅")
    else:
        print("ERROR: 必须 --inject <内容JSON> 或 --from-recommend（本项目内容由 WorkBuddy 模型注入，不调用 GLM）")
        sys.exit(1)

    print("\n小红书成稿+生图完成 ✅（下一步：push_to_feishu_content.py 推飞书；或 xiaohongshu-mcp 推草稿箱）")


if __name__ == "__main__":
    main()
