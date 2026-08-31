# -*- coding: utf-8 -*-
"""
topic_pool.py —— 企业级选题池（human-in-the-loop 闸门，极光AIGC · 免费本地版）

设计依据：research/企业级竞品爬取与选题方案_2026-08-31.md §3
- 替代 plan_of_the_day.py 里的 GLM 调用：推荐理由用规则化生成（热度/收藏率/评论率/搜索词/增速），
  不调任何外部 LLM；内容生成由运营/AI 在「拍板」后注入（见 xhs_mvp.py --inject）。
- 关键设计：Agent 只把信号整理成「可判断事件 + 推荐理由 + 指数」，[不替人下结论]；
  运营在飞书「选题池」里点 1 个或说「换一批」，**拍板后才触发文案生成**（闭合用户「没选好就做=浪费资源」反馈）。

对外能力：
  - build_topic_pool(scored)  把 analytics 打分的 notes 转成 选题池 行（含推荐理由 + 热度指数 + 优先级 + 状态=待拍板）
  - recommend_reason(note)    规则化中文推荐理由（搜索价值 + 共鸣价值 + 时效 + 721 切入建议）
  - write_topic_pool_xlsx()   写 pipeline/选题池.xlsx（飞书「选题池」数据源）
  - write_today_recommend()   写 pipeline/今日选题推荐.xlsx（下游 xhs_mvp --from-recommend 用）
  - pick(topics)              人类拍板：返回被选中的候选（仅此处才进入成稿）
CLI: python topic_pool.py --csv <file> --top 10
"""
import os, sys, json, argparse
from openpyxl import Workbook, load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, HERE)
import analytics as A

POOL_XLSX = os.path.join(HERE, "选题池.xlsx")
TODAY_XLSX = os.path.join(HERE, "今日选题推荐.xlsx")
POOL_HEADERS = ["排名", "选题标题", "对标爆款", "推荐理由", "热度指数", "优先级", "状态", "内容方向"]
TODAY_HEADERS = ["排名", "选题标题", "对标爆款", "内容方向", "为什么发", "配图提示", "状态"]


def _dedup_by_note(scored):
    best = {}
    for n in scored:
        nid = n.get("note_id")
        if not nid:
            continue
        if nid not in best or n["heat_score"] > best[nid]["heat_score"]:
            best[nid] = n
    return list(best.values())


def recommend_reason(note, total):
    """规则化推荐理由（人话、运营小白可懂）。"""
    rank = note.get("heat_rank", "?")
    heat = note.get("heat_score", 0)
    r = note.get("R_total", 0)
    cr = note.get("collect_rate", 0)
    cmr = note.get("comment_rate", 0)
    vel = note.get("velocity_per_day", 0)
    src = note.get("source") or note.get("source_keyword") or "竞品赛道"
    pct = round(100 - (rank - 1) / max(total, 1) * 100) if total else 0

    parts = []
    # 1) 热度（相对赛道）
    parts.append(f"热度分 {heat}（赛道前 {pct}%，相对同赛道中位数 {r}×）——说明这个方向的内容在同赛道表现突出。")
    # 2) 有用性（收藏=想存下来用）
    if cr >= 0.5:
        parts.append(f"收藏率 {cr:.0%} 偏高，用户「想存下来用」的信号强，适合做成教程/清单体。")
    # 3) 话题性（评论=共鸣/争议）
    if cmr >= 0.005:
        parts.append(f"评论率 {cmr:.1%} 不低，有讨论/共鸣，开头可用钩子抛痛点。")
    # 4) 搜索价值（来源搜索词 = 用户主动在搜）
    parts.append(f"来源搜索词「{src}」——用户在主动搜这类内容（搜索流量占比 50-65%，搜索导向更易拿自然流量）。")
    # 5) 时效
    if vel >= 500:
        parts.append(f"互动速度 {vel:.0f}/天，正在起量，建议尽快跟。")
    else:
        parts.append(f"互动速度 {vel:.0f}/天，长尾平稳，可排期做。")
    # 6) 721 切入建议（微创新，不抄）
    parts.append("切入建议（721）：把爆款视角换成你「医疗AI/医学生」人设能真做出来的版本（微创新 70%），别直接抄。")
    return "\n".join(parts)


def build_topic_pool(scored, top_n=10):
    """scored: analytics.score_notes 的输出（已含 heat_score 等）。返回 选题池 行 list。"""
    scored = _dedup_by_note(scored)
    scored.sort(key=lambda x: x["heat_score"], reverse=True)
    total = len(scored)
    rows = []
    for i, n in enumerate(scored[:top_n], 1):
        reason = recommend_reason(n, total)
        rows.append({
            "排名": i,
            "选题标题": n.get("title", "")[:40],
            "对标爆款": f"{n.get('account','?')}《{n.get('title','')[:20]}》赞{n.get('liked',0)}藏{n.get('collected',0)}评{n.get('comment',0)}",
            "推荐理由": reason,
            "热度指数": n.get("heat_score", 0),
            "优先级": i,
            "状态": "待拍板",
            "内容方向(拍板后填)": "",
        })
    return rows


def write_topic_pool_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "选题池"
    ws.append(POOL_HEADERS)
    for r in rows:
        ws.append([r.get(h, "") for h in POOL_HEADERS])
    wb.save(POOL_XLSX)
    return POOL_XLSX


def write_today_recommend(rows):
    """下游 xhs_mvp --from-recommend 需要 选题标题 列。状态标 待拍板，拍板前不生成。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "今日选题推荐"
    ws.append(TODAY_HEADERS)
    for r in rows:
        ws.append([r.get("排名", ""), r.get("选题标题", ""), r.get("对标爆款", ""),
                   "（待拍板后由 AI 生成）", r.get("推荐理由", ""),
                   "（待拍板后由 AI 生成）", "待拍板"])
    wb.save(TODAY_XLSX)
    return TODAY_XLSX


def pick(rows, ranks):
    """人类拍板：返回选中的候选（仅这些进入成稿）。ranks 为排名列表。"""
    chosen = [r for r in rows if r["排名"] in ranks]
    return chosen


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv")
    ap.add_argument("--xlsx")
    ap.add_argument("--top", type=int, default=10)
    a = ap.parse_args()
    if a.csv:
        notes = A.load_notes_from_csv(a.csv)
    elif a.xlsx:
        notes = A.load_notes_from_xlsx(a.xlsx)
    else:
        print("ERROR: 需 --csv 或 --xlsx"); sys.exit(1)
    scored, _ = A.score_notes(notes)
    rows = build_topic_pool(scored, a.top)
    p1 = write_topic_pool_xlsx(rows)
    p2 = write_today_recommend(rows)
    p3 = A.write_heat_board_xlsx(scored)
    print(f"✅ 选题池写好了（{len(rows)} 条，全部状态=待拍板）：")
    print(f"   {p1}")
    print(f"   {p2}")
    print(f"   热度看板：{p3}")
    print("\n—— 运营在飞书「选题池」看推荐理由+指数，说「用第 N 条」即拍板，才进成稿（human-in-the-loop 闸门）——")
    for r in rows[:5]:
        print(f"  {r['排名']}. [{r['热度指数']}] {r['选题标题']}")
        print(f"     └ {r['推荐理由'].splitlines()[0]}")


if __name__ == "__main__":
    main()
